#!/usr/bin/env python3
"""Generate the Q2 2026 (July) US Filings Calendar tab.

Mirrors the Q1 2026 April calendar layout used in the 2026 Filings Calendar
spreadsheet. Q2 quarter-end is June 30, 2026; July is the primary filing month.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

OUTPUT = Path(__file__).resolve().parent.parent / "output" / "Q2_2026_July_Filings_Calendar.xlsx"

Q2_PLAN_URL = (
    "https://rippling.atlassian.net/wiki/spaces/ENG/pages/6193446978/"
    "TE+-+Q2+2026+QE+Planning+Tracker"
)

# day keys 1-7 = Sunday through Saturday within each week block
WEEKS: list[tuple[list[str], list[tuple[str, dict[int, str]]]]] = [
    (
        ["Jun 28", "Jun 29", "Jun 30", "1", "2", "3", "4"],
        [
            (
                "Semi-Monthly Filings / Payments",
                {
                    4: (
                        "-Eng: SM Filings Process Kickoff & FF File Generation Handover\n"
                        "All paper files must be approved for 9:30am PST for paper packet "
                        "to hit Lob's API by 10am PST"
                    ),
                },
            ),
            (
                "Monthly Filings / Payments",
                {4: "Eng: Monthly Filings Process Kickoff"},
            ),
            (
                "Quarterly Filings / Payments",
                {
                    3: (
                        "Continuation of US Filings Address Pre-Audit (Ops)\n"
                        "Continuation of HAB Id Pre-Audit (Ops)\n"
                        "Data Exchange Deadline (Ops)"
                    ),
                    4: (
                        "Close of Quarter (Q2 2026)\n"
                        "Last day to accept Account ID / Rate changes\n"
                        "Eng: QE Filings Process Kickoff"
                    ),
                },
            ),
            (
                "Second Filings Batch",
                {
                    4: (
                        "Deadline to provide the initial estimated list of clients with "
                        "missing account number + missing tax rates"
                    ),
                },
            ),
            (
                "Filings / Payments Audits",
                {4: "Complete audit for semi-monthly filings due on the 3rd"},
            ),
            (
                "Client Communications / Recon",
                {
                    4: (
                        "QWR / Recon Run Initiated (est.)\n"
                        "NY MCTMT: Tax due calculations completed"
                    ),
                    5: (
                        "NY MCTMT: Payment to NY initiated\n"
                        "Client list with MCTMT debit values needed"
                    ),
                    6: (
                        "NY MCTMT: Payment received by NY\n"
                        "Client debit comms sent"
                    ),
                },
            ),
        ],
    ),
    (
        ["5", "6", "7", "8", "9", "10", "11"],
        [
            ("Semi-Monthly Filings / Payments", {}),
            (
                "Monthly Filings / Payments",
                {
                    3: "Ideal target for monthly file handover",
                    4: "Begin submission for all monthly filings bulk files",
                    5: (
                        "Deadline to send KYNELSONCOFILE paper filings to Lob by 9:30am PST"
                    ),
                    7: (
                        "Monthly filings deadline (10th)\n"
                        "Target for quarterly file handover\n"
                        "Begin monthly file review and approvals"
                    ),
                },
            ),
            (
                "Quarterly Filings / Payments",
                {3: "Eng: QE Filings Process\nSecond Filings Batch"},
            ),
            (
                "Filings / Payments Audits",
                {
                    4: "Complete audit for monthly filings due on the 10th",
                    7: "Final verification of 10th deadline agencies",
                },
            ),
            (
                "Client Communications / Recon",
                {6: "NY MCTMT Client Debit (July 9)"},
            ),
        ],
    ),
    (
        ["12", "13", "14", "15", "16", "17", "18"],
        [
            (
                "Semi-Monthly Filings / Payments",
                {
                    2: (
                        "Begin quarterly file review and approval + Wrap up monthly bulk filings"
                    ),
                    5: (
                        "-Eng: SM Filings Process Kickoff & FF File Generation Handover\n"
                        "All paper files must be approved for 9:30am PST for paper packet "
                        "to hit Lob's API by 10am PST"
                    ),
                },
            ),
            (
                "Monthly Filings / Payments",
                {
                    3: (
                        "All paper filings due on the 15th must be approved for 9:30am PST "
                        "for paper packet to hit Lob's API by 10am PST"
                    ),
                    4: "Monthly filings deadline (15th)",
                    5: (
                        "All paper filings due on the 20th must be approved for 9:30am PST "
                        "for paper packet to hit Lob's API by 10am PST"
                    ),
                },
            ),
            (
                "Quarterly Filings / Payments",
                {
                    2: "Begin submission for all quarterly filings bulk files",
                    4: (
                        "Quarterly filings deadline for the following agencies:\n"
                        "-HISWFILE\n"
                        "-MSSWFILE\n"
                        "-MDSW (manual $0 hand-key)\n"
                        "-DELOCALWILMINGTONFILE (Paper)"
                    ),
                    5: (
                        "Begin submission for all quarterly filings bulk files\n"
                        "1st Recon Customer Debits\n"
                        "Client Comms auto-triggered"
                    ),
                },
            ),
            (
                "Filings / Payments Audits",
                {
                    4: (
                        "Complete audit for monthly and quarterly filings due on the 15th"
                    ),
                    5: "Complete audit for semi-monthly filings due on the 18th/20th",
                    6: (
                        "Complete audit for monthly and quarterly filings due on the 20th "
                        "& semi-monthly filings due on the 18th/20th"
                    ),
                },
            ),
            (
                "Second Filings Batch",
                {
                    5: "Deadline for agency account inputs for PEO Ops/Customers",
                    6: (
                        "Comms for the 1st recon debits will go out in batches between "
                        "the 15th and 17th"
                    ),
                },
            ),
        ],
    ),
    (
        ["19", "20", "21", "22", "23", "24", "25"],
        [
            (
                "Monthly Filings / Payments",
                {
                    2: "Monthly filings deadline (20th)",
                    4: (
                        "All paper filings due on the 25th must be approved for 9:30am PST "
                        "for paper packet to hit Lob's API by 10am PST"
                    ),
                    6: "Final verification of 25th deadline agencies",
                    7: (
                        "Monthly/Quarterly deadline falls on weekend, 25th deadline shifts to 27th"
                    ),
                },
            ),
            (
                "Quarterly Filings / Payments",
                {1: "Quarterly filings deadline"},
            ),
            (
                "Filings / Payments Audits",
                {4: "Complete audit for monthly and quarterly filings due on the 25th"},
            ),
            (
                "Second Filings Batch",
                {
                    2: "Engg to share the Impact List for 2nd batch with TaxOps",
                    4: "Recon to be completed for 2nd batch",
                    6: "2nd recon debits to be completed by this date",
                },
            ),
        ],
    ),
    (
        ["26", "27", "28", "29", "30", "31", "Aug 1"],
        [
            (
                "Monthly Filings / Payments",
                {
                    2: "Monthly filings deadline",
                    3: (
                        "All paper filings due on the 30th/EOM must be approved for 9:30am PST "
                        "for paper packet to hit Lob's API by 10am PST"
                    ),
                    4: "Final verification of EOM deadline agencies",
                    5: "Monthly filings deadline (EOM)",
                },
            ),
            (
                "Quarterly Filings / Payments",
                {
                    2: "Quarterly filings deadline",
                    5: "Quarterly filings EOM deadline",
                    6: (
                        "FEDERAL Q2 941 DEADLINE (July 31, 2026)\n"
                        "State SUI Q2 Filing Deadline (July 31, 2026)"
                    ),
                },
            ),
            (
                "Filings / Payments Audits",
                {
                    3: "Complete audit for monthly and quarterly filings due on the 30th/EOM",
                    5: "Final verification of EOM deadline agencies in second batch",
                },
            ),
            (
                "Second Filings Batch",
                {2: "Kick off 2nd Batch File Submissions"},
            ),
            (
                "Client Communications / Recon",
                {
                    3: "Share ACH Debit Payment Files to Payments Eng for 2nd Recon",
                    6: (
                        "QE Packages Generation (Target: ~July 31)\n"
                        "QE Recon Package available to clients (~5 days before standard recon debit)\n"
                        "Pre-QE Notification: agency-by-agency listing of liabilities, amounts "
                        "funded, payments made by Rippling, and variance\n"
                        "Notification via in-product, mobile push, and email"
                    ),
                },
            ),
        ],
    ),
]


def _style_header_row(ws, row: int) -> None:
    header_fill = PatternFill("solid", fgColor="D9E1F2")
    bold = Font(bold=True)
    for col in range(1, 9):
        cell = ws.cell(row=row, column=col)
        cell.font = bold
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical="top")


def _write_week(
    ws,
    start_row: int,
    dates: list[str],
    categories: list[tuple[str, dict[int, str]]],
    *,
    include_resources: bool = False,
) -> int:
    row = start_row
    if include_resources:
        ws.cell(row=row, column=1, value="July")
        for idx, title in enumerate(
            ["Sunday", "Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"],
            start=2,
        ):
            ws.cell(row=row, column=idx, value=title)
        ws.cell(row=row, column=9, value="Important Resources")
        _style_header_row(ws, row)
        row += 1

    for idx, day in enumerate(dates, start=2):
        ws.cell(row=row, column=idx, value=day)
    if include_resources:
        ws.cell(row=row, column=9, value="Q2 2026 Filings Calendar Plan (Confluence)")
        ws.cell(row=row, column=10, value=Q2_PLAN_URL)
    row += 1

    for category, day_notes in categories:
        ws.cell(row=row, column=1, value=category)
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True, vertical="top")
        for day_idx, text in day_notes.items():
            cell = ws.cell(row=row, column=day_idx + 1, value=text)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        row += 1

    row += 1
    return row


def build_workbook() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "July"

    row = 1
    for week_idx, (dates, categories) in enumerate(WEEKS):
        row = _write_week(
            ws,
            row,
            dates,
            categories,
            include_resources=week_idx == 0,
        )

    widths = [34, 18, 18, 18, 22, 22, 22, 18, 28, 60]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    return wb


def main() -> None:
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb = build_workbook()
    wb.save(OUTPUT)
    print(f"Wrote {OUTPUT}")


if __name__ == "__main__":
    main()
